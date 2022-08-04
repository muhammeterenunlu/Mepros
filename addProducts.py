import requests
from os.path import basename
import os
from bs4 import BeautifulSoup
import openpyxl

workbookObj = openpyxl.load_workbook(os.path.expanduser('~/Downloads/meprosProductCodes.xlsx'))
sheetObj = workbookObj.active;

def downloadImages(soup, productName):
    images = soup.find_all(class_='prettyphoto')

    i = 0
    for image in images:
        imageContent = image['href']

        fileName = 'lighting-images/{}{}.png'.format(
                                            productName.upper(), 
                                            '-1' if i > 0 else '')
        
        with open(fileName, 'wb') as f:
            f.write(requests.get(imageContent).content)
        i+=1

def createData(productId):
    productSearchId = f'mps-{productId}'

    productData = f'''\
mps_{productId}_data = {{
  name: "mps-{productId}",
  colorsImagesDictionary: {{
    "#ffffff": [
      "../lighting-images/MPS{productId.upper()}.png",
      "../lighting-images/MPS{productId.upper()}-1.png",
    ],
  }},
  tags: [
    "Material: Metal Aluminum Body, electrostatic powder coat",
    "Socket: e27",
    "Volt: 220V",
  ],
}}

globals.productManager.addProduct(new LightingProduct(mps_{productId}_data));'''

    fileHandle = open(f'ProductDatas/Lighting/mps_{productId}_data.js', 'x')
    fileHandle.write(productData)

def addToHtml(productId):
    fileHandle = open('ProductPageTemplates/lighting2ImagePage.html', 'r+')

    lineToBeInserted = f'  <script defer src="../ProductDatas/Lighting/mps_{productId}_data.js"></script>\n'
    
    lines = fileHandle.readlines()
    lines.insert(80, lineToBeInserted)

    fileHandle.seek(0)
    for line in lines:
        fileHandle.write(line)

for i in range(106, 141):

    productUrlCell = sheetObj.cell(row = i, column = 3)
    productNameCell = sheetObj.cell(row = i, column = 2)

    response = requests.get(productUrlCell.value)

    soup = BeautifulSoup(response.content, 'html.parser')


    productName = productNameCell.value.lower()
    productId = productName[3:]

    downloadImages(soup, productName)
    createData(productId)
    addToHtml(productId)

